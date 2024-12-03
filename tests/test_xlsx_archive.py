from openpyxl import load_workbook
import zipfile
from files_pathes import *


def test_xlsx_to_zip(create_archive):
    with zipfile.ZipFile(ARCHIVE_PATH) as x_f:
        xlsx_archived = x_f.extract('file_example_XLSX_50.xlsx')
        workbook = load_workbook(xlsx_archived)
        sheet = workbook.active
        assert sheet.cell(row=49, column=2).value == 'Demetria'
        os.remove('file_example_XLSX_50.xlsx')

        assert os.path.exists(ARCHIVE_PATH)


def test_archive_deleted(create_archive):
    os.remove(ARCHIVE_PATH)
    assert not os.path.exists(ARCHIVE_PATH)

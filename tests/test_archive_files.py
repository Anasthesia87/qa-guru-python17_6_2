import csv
from io import StringIO, BytesIO
from pypdf import PdfReader
from openpyxl import load_workbook
import zipfile
from files_paths import *


def test_csv_to_zip(create_archive):
    with zipfile.ZipFile(ARCHIVE_PATH) as csv_file:
        with csv_file.open('user_file.csv') as file_example_csv:
            content = file_example_csv.read().decode('utf-8')

            file_like_object = StringIO(content)
            csv_rows = csv.reader(file_like_object)
            new_list = []
            for row in csv_rows:
                new_list.append(row)

            assert len(new_list) == 3, "Количество строк в выбранном файле не соответствует ожидаемому."
            assert new_list == [['1', 'alice', 'alice@example.com'],
                                ['2', 'bob', 'bob@example.com'],
                                ['3', 'carol',
                                 'carol@example.com']], "Содержание выбранного файла не совпадает с ожидаемым."


def test_pdf_to_zip(create_archive):
    with zipfile.ZipFile(ARCHIVE_PATH) as pdf_file:
        with pdf_file.open('Python Testing with Pytest (Brian Okken).pdf') as file_example_pdf:
            content = BytesIO(file_example_pdf.read())

        reader = PdfReader(content)
        assert len(reader.pages) == 256, "Количество страниц в выбранном файле не соответствует ожидаемому."
        page = reader.pages[1]
        text = page.extract_text()
        print(text)

        assert 'Copyright © 2017 The Pragmatic Programmers, LLC' in text, "Содержание выбранного файла не совпадает с ожидаемым."


def test_xlsx_to_zip(create_archive):
    with zipfile.ZipFile(ARCHIVE_PATH) as xlsx_file:
        with xlsx_file.open('file_example_XLSX_50.xlsx') as file_example_xlsx:
            content = BytesIO(file_example_xlsx.read())

        workbook = load_workbook(filename=content)
        sheet = workbook.active

        assert sheet.cell(row=49,
                          column=2).value == 'Demetria', "Значение в указанной ячейке не соответствует ожидаемому."
